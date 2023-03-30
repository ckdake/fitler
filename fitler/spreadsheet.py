from fitler.metadata import ActivityMetadata

import openpyxl
from pathlib import Path

import json
from dateutil import parser as dateparser

class ActivitySpreadsheet(object):
    def __init__(self, path):
        self.path = path
        self.activities_metadata = []

    def parse(self):
        xlsx_file = Path('ActivityData', self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        col_names = []
        for column in sheet.iter_cols(1, sheet.max_column):
            col_names.append(column[0].value)

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                am_dict = {}
                am_dict['date'] = dateparser.parse(str(row[0])).strftime("%Y-%m-%d")
                if activity_type := row[1]: am_dict['activity_type'] = activity_type
                if location_name := row[2]: am_dict['location_name'] = location_name
                if city := row[3]: am_dict['city'] = city
                if state := row[4]: am_dict['state'] = state
                if temperature := row[5]: am_dict['temperature'] = temperature
                if equipment := row[6]: am_dict['equipment'] = equipment
                if duration_hms := row[7]: am_dict['duration_hms'] = duration_hms
                # row[8] is calculated 'duration_h', 
                if distance := row[9]: am_dict['distance'] = distance
                if max_speed := row[10]: am_dict['max_speed'] = max_speed
                if avg_heart_rate := row[11]: am_dict['avg_heart_rate'] = avg_heart_rate
                if max_heart_rate := row[12]: am_dict['max_heart_rate'] = max_heart_rate
                if calories := row[13]: am_dict['calories'] = calories
                if max_elevation := row[14]: am_dict['max_elevation'] = max_elevation
                if total_elevation_gain := row[15]: am_dict['total_elevation_gain'] = total_elevation_gain
                if with_names := row[16]: am_dict['with_names'] = with_names
                if avg_cadence := row[17]: am_dict['avg_cadence'] = avg_cadence
                if strava_id := row[18]: am_dict['strava_id'] = strava_id
                if garmin_id := row[19]: am_dict['garmin_id'] = garmin_id
                if ridewithgps_id := row[20]: am_dict['ridewithgps_id'] = ridewithgps_id
                if notes := row[21]: am_dict['notes'] = notes

                am_dict['source'] = 'Spreadsheet'
                am, created = ActivityMetadata.get_or_create(**am_dict)
                am.save()

                self.activities_metadata.append(am)
    
    def to_json(self):
        return json.dumps(self, default=lambda o: o.__dict__, sort_keys=True, indent=4) 