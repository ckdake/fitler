"""
spreadsheet_activity_collection.py: a spreadsheet of activities
"""

import os
import openpyxl

from fitler.activity_collection import ActivityCollection
from fitler.spreadsheet_activity import SpreadsheetActivity


class SpreadsheetActivityCollection(ActivityCollection):
    """A standardized object reflecting a fitness activity"""

    def __init__(self, path=None):
        self.path = path
        self.activities = []

    def initialize(self) -> bool:
        xlsx_file = os.path.abspath(self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        col_names = []
        for column in sheet.iter_cols(1, sheet.max_column):
            col_names.append(column[0].value)

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                self.activities.append(SpreadsheetActivity(row, i))
        return True
