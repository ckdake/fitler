"""Spreadsheet provider for Fitler.

This module defines the SpreadsheetActivities class, which provides an interface
for interacting with activity data stored in local spreadsheet files.
"""

from typing import List, Dict, Any, Optional
from pathlib import Path
import openpyxl
from dateutil import parser as dateparser
import datetime

from fitler.providers.base import FitnessProvider, Activity


class SpreadsheetActivities(FitnessProvider):
    def __init__(self, path: str):
        self.path = path

    def _seconds_to_hms(self, seconds: Optional[float]) -> str:
        if seconds is None:
            return ""
        try:
            seconds = int(round(seconds))
            return str(datetime.timedelta(seconds=seconds))
        except Exception:
            return ""

    def _hms_to_seconds(self, hms: Optional[str]) -> Optional[float]:
        if not hms:
            return None
        try:
            t = datetime.datetime.strptime(hms, "%H:%M:%S")
            return t.hour * 3600 + t.minute * 60 + t.second
        except Exception:
            try:
                # Try MM:SS
                t = datetime.datetime.strptime(hms, "%M:%S")
                return t.minute * 60 + t.second
            except Exception:
                return None

    def fetch_activities(self) -> List[Activity]:
        xlsx_file = Path("ActivityData", self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        if sheet is None:
            return []

        activities: List[Activity] = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue  # Skip header row
            activity_kwargs: Dict[str, Any] = {}
            activity_kwargs["start_time"] = dateparser.parse(str(row[0])).strftime(
                "%Y-%m-%d"
            )
            if row[1]:
                activity_kwargs["activity_type"] = row[1]
            if row[2]:
                activity_kwargs["location_name"] = row[2]
            if row[3]:
                activity_kwargs["city"] = row[3]
            if row[4]:
                activity_kwargs["state"] = row[4]
            if row[5]:
                activity_kwargs["temperature"] = row[5]
            if row[6]:
                activity_kwargs["equipment"] = row[6]
            # duration_hms is stored as HH:MM:SS, but we want .duration in seconds
            if row[7]:
                activity_kwargs["duration"] = self._hms_to_seconds(row[7])
            if row[8]:
                activity_kwargs["distance"] = row[8]
            if row[9]:
                activity_kwargs["max_speed"] = row[9]
            if row[10]:
                activity_kwargs["avg_heart_rate"] = row[10]
            if row[11]:
                activity_kwargs["max_heart_rate"] = row[11]
            if row[12]:
                activity_kwargs["calories"] = row[12]
            if row[13]:
                activity_kwargs["max_elevation"] = row[13]
            if row[14]:
                activity_kwargs["total_elevation_gain"] = row[14]
            if row[15]:
                activity_kwargs["with_names"] = row[15]
            if row[16]:
                activity_kwargs["avg_cadence"] = row[16]
            if row[17]:
                activity_kwargs["strava_id"] = row[17]
            if row[18]:
                activity_kwargs["garmin_id"] = row[18]
            if row[19]:
                activity_kwargs["ridewithgps_id"] = row[19]
            if row[20]:
                activity_kwargs["notes"] = row[20]
            activity_kwargs["source_file"] = str(xlsx_file)
            activity_kwargs["source_file_type"] = "spreadsheet"
            activity_kwargs["spreadsheet_id"] = i + 1
            activities.append(Activity(**activity_kwargs))
        return activities

    def get_activity_by_id(self, activity_id: str) -> Optional[Activity]:
        xlsx_file = Path("ActivityData", self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        row_idx = int(activity_id)
        if row_idx <= 1 or row_idx > sheet.max_row:
            return None

        row = [cell.value for cell in sheet[row_idx]]
        activity_kwargs: Dict[str, Any] = {}
        activity_kwargs["start_time"] = (
            dateparser.parse(str(row[0])).strftime("%Y-%m-%d") if row[0] else ""
        )
        if row[1]:
            activity_kwargs["activity_type"] = row[1]
        if row[2]:
            activity_kwargs["location_name"] = row[2]
        if row[3]:
            activity_kwargs["city"] = row[3]
        if row[4]:
            activity_kwargs["state"] = row[4]
        if row[5]:
            activity_kwargs["temperature"] = row[5]
        if row[6]:
            activity_kwargs["equipment"] = row[6]
        if row[7]:
            activity_kwargs["duration"] = self._hms_to_seconds(row[7])
        if row[8]:
            activity_kwargs["distance"] = row[8]
        if row[9]:
            activity_kwargs["max_speed"] = row[9]
        if row[10]:
            activity_kwargs["avg_heart_rate"] = row[10]
        if row[11]:
            activity_kwargs["max_heart_rate"] = row[11]
        if row[12]:
            activity_kwargs["calories"] = row[12]
        if row[13]:
            activity_kwargs["max_elevation"] = row[13]
        if row[14]:
            activity_kwargs["total_elevation_gain"] = row[14]
        if row[15]:
            activity_kwargs["with_names"] = row[15]
        if row[16]:
            activity_kwargs["avg_cadence"] = row[16]
        if row[17]:
            activity_kwargs["strava_id"] = row[17]
        if row[18]:
            activity_kwargs["garmin_id"] = row[18]
        if row[19]:
            activity_kwargs["ridewithgps_id"] = row[19]
        if row[20]:
            activity_kwargs["notes"] = row[20]
        activity_kwargs["source_file"] = str(xlsx_file)
        activity_kwargs["source_file_type"] = "spreadsheet"
        activity_kwargs["spreadsheet_id"] = row_idx
        return Activity(**activity_kwargs)

    def update_activity(self, activity_id: str, activity: Activity) -> bool:
        xlsx_file = Path("ActivityData", self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        row_idx = int(activity_id)
        if row_idx <= 1 or row_idx > sheet.max_row:
            return False

        # Prepare the row in the same order as the header
        row = [
            getattr(activity, "start_time", ""),
            getattr(activity, "activity_type", ""),
            getattr(activity, "location_name", ""),
            getattr(activity, "city", ""),
            getattr(activity, "state", ""),
            getattr(activity, "temperature", ""),
            getattr(activity, "equipment", ""),
            self._seconds_to_hms(getattr(activity, "duration", None)),
            getattr(activity, "distance", ""),
            getattr(activity, "max_speed", ""),
            getattr(activity, "avg_heart_rate", ""),
            getattr(activity, "max_heart_rate", ""),
            getattr(activity, "calories", ""),
            getattr(activity, "max_elevation", ""),
            getattr(activity, "total_elevation_gain", ""),
            getattr(activity, "with_names", ""),
            getattr(activity, "avg_cadence", ""),
            getattr(activity, "strava_id", ""),
            getattr(activity, "garmin_id", ""),
            getattr(activity, "ridewithgps_id", ""),
            getattr(activity, "notes", ""),
        ]
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
        wb_obj.save(xlsx_file)
        return True

    def get_gear(self) -> Dict[str, str]:
        xlsx_file = Path("ActivityData", self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        gear_set = set()
        # The equipment column is index 6 (0-based)
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue  # Skip header row
            equipment = row[6]
            if equipment:
                gear_set.add(str(equipment))
        # Use the equipment name as both key and value
        return {name: name for name in gear_set}

    def create_activity(self, activity: Activity) -> str:
        xlsx_file = Path("ActivityData", self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        row = [
            getattr(activity, "start_time", ""),
            getattr(activity, "activity_type", ""),
            getattr(activity, "location_name", ""),
            getattr(activity, "city", ""),
            getattr(activity, "state", ""),
            getattr(activity, "temperature", ""),
            getattr(activity, "equipment", ""),
            self._seconds_to_hms(getattr(activity, "duration", None)),
            getattr(activity, "distance", ""),
            getattr(activity, "max_speed", ""),
            getattr(activity, "avg_heart_rate", ""),
            getattr(activity, "max_heart_rate", ""),
            getattr(activity, "calories", ""),
            getattr(activity, "max_elevation", ""),
            getattr(activity, "total_elevation_gain", ""),
            getattr(activity, "with_names", ""),
            getattr(activity, "avg_cadence", ""),
            getattr(activity, "strava_id", ""),
            getattr(activity, "garmin_id", ""),
            getattr(activity, "ridewithgps_id", ""),
            getattr(activity, "notes", ""),
        ]
        sheet.append(row)
        wb_obj.save(xlsx_file)
        return str(sheet.max_row)

    def set_gear(self, gear_id: str, activity_id: str) -> bool:
        xlsx_file = Path("ActivityData", self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        row_idx = int(activity_id)
        # The gear column is index 7 (0-based), so column 7+1=8 (1-based for openpyxl)
        gear_col = 7 + 1
        if row_idx <= 1 or row_idx > sheet.max_row:
            return False  # Invalid row (header or out of range)
        sheet.cell(row=row_idx, column=gear_col, value=gear_id)
        wb_obj.save(xlsx_file)
        return True
